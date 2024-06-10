from xml.dom import minidom
import os
import openpyxl
import win32clipboard

numero_nfs = {"MAVIFER": [], "IZAMAC": []}


def main():
    total = load_xml()

    if len(itens := check_duplicate(total["MAVIFER"], total["IZAMAC"])) > 0:
        for item in itens:
            while True:
                empresa = input(f"{item} é de qual empresa (mavifer ou izamac)? ").upper()
                if empresa in ["MAVIFER", "IZAMAC"]:
                    if empresa == "MAVIFER":
                        value = total["IZAMAC"].pop(item)
                        total[empresa][item] += value
                    elif empresa == "IZAMAC":
                        value = total["MAVIFER"].pop(item)
                        total[empresa][item] += value
                    break

    faturamento()
    while True:
        print("1. Apontamento de Expedicao")
        print("2. Romaneio")
        print("0. Sair")
        opcao = int(input())
        if opcao in [1, 2, 0]:
            break

    match opcao:
        case 1:
            apontamento_expedicao(total)
        case 2:
            romaneio(total)
        case 0:
            return


# le os arquivos xml e retorna um dicionario com os itens e quantidades desses xml
def load_xml() -> dict:
    xml_path = "C:\\Users\\Nota Fiscal\\Documents\\GitHub\\auto-work\\xml\\"
    total = {"MAVIFER": {}, "IZAMAC": {}}
    for file_name in os.listdir(xml_path):
        nfe = minidom.parse(xml_path + file_name)
        nnf = nfe.getElementsByTagName("nNF")[0].childNodes[0].nodeValue
        empresa = nfe.getElementsByTagName("xNome")[0].childNodes[0].nodeValue.split(" ")[0].strip()
        itens = nfe.getElementsByTagName("prod")
        cfop = itens[0].getElementsByTagName("CFOP")[0].childNodes[0].nodeValue
        # print(f"Empresa: {empresa} | NF {nnf} | CFOP: {cfop}")
        numero_nfs[empresa].append(nnf)

        if cfop == "5124" or cfop == "5916":
            for item in itens:
                cod_prod = item.getElementsByTagName("cProd")[0].childNodes[0].nodeValue.strip()
                qtd_prod = item.getElementsByTagName("qCom")[0].childNodes[0].nodeValue
                if cod_prod in total[empresa]:
                    total[empresa][cod_prod] += float(qtd_prod)
                else:
                    total[empresa][cod_prod] = float(qtd_prod)
    return total


# recebe dois dicionarios e verifica se há algum item duplicado e retorna uma lista com os itens duplicados
def check_duplicate(mavifer: dict, izamac: dict) -> list:
    duplicate = []
    for key in mavifer.keys():
        if key in izamac.keys():
            duplicate.append(key)

    return duplicate


# recebe uma tabela e retorna uma lista com os itens já estão na planilha
def get_items_excel(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict:
    items = {}
    for cell in ws['A6':'A35']:
        if cell[0].value is not None:
            items[cell[0].value.upper()] = cell[0].row
    return items


# Retorna um inteiro indicando qual a proxima linha da tabela que esta vazia
def get_first_empty(items: dict) -> int:
    row = 5
    for item in items.keys():
        if items[item] > row:
            row = items[item]
    return row + 1


# recebe um dicionario com os itens das duas empresas e salva na planilha
def apontamento_expedicao(total: dict):
    janelas = {
        1: 'B',
        2: 'C',
        3: 'D',
        4: 'E',
        5: 'F',
        6: 'G',
        7: 'H',
        8: 'I',
        9: 'J',
        10: 'K'
    }

    while True:
        num_janela = int(input("Qual janela da carga?(1 .. 10) ")) + 1
        if num_janela in janelas:
            break

    path = "\\\\10.1.1.7\\Documents\\PRIORIDADES\\Apontamento Expedição.xlsx"
    wb = openpyxl.load_workbook(path)

    # range header B4:K4
    # range items A6:A35
    # range qtd B6:K35
    # col = B:K
    # row = 6:35

    for emp_name in ["MAVIFER", "IZAMAC"]:
        ws = wb[emp_name]

        for item in total[emp_name].keys():
            items = get_items_excel(ws)
            if item in items.keys():
                cell_coord = f"{janelas[num_janela]}{items[item]}"
                ws[cell_coord] = total[emp_name][item]
            else:
                row = get_first_empty(items)
                cell_coord = f"{janelas[num_janela]}{row}"
                ws[f"A{row}"] = item
                ws[cell_coord] = total[emp_name][item]

    wb.save("\\\\10.1.1.7\\Documents\\PRIORIDADES\\Apontamento Expedição.xlsx")


def faturamento():
    xml_path = "C:\\Users\\Nota Fiscal\\Documents\\GitHub\\auto-work\\xml\\"

    cfop_map = {
        "5916": "Conserto",
        "5902": "R. Industrialização",
        "5124": "Valor\t\tPeso",
        "5921": "Embalagem\tColar Schulz",
        "5949": "Simples Remessa\tResiduo de Ferro"
    }

    final_text = ""

    for file_name in os.listdir(xml_path):
        nfe = minidom.parse(xml_path + file_name)
        nnf = nfe.getElementsByTagName("nNF")[0].childNodes[0].nodeValue
        # empresa = nfe.getElementsByTagName("xNome")[0].childNodes[0].nodeValue.split(" ")[0].strip()
        itens = nfe.getElementsByTagName("prod")
        cfop = itens[0].getElementsByTagName("CFOP")[0].childNodes[0].nodeValue

        if cfop == "5124":
            valor = nfe.getElementsByTagName("vLiq")[0].childNodes[0].nodeValue
            peso = str(round(float(nfe.getElementsByTagName("pesoL")[0].childNodes[0].nodeValue) / 1000, 1))
            cfop_map[cfop] = f"{valor.replace('.', ',')}\t\t{peso.replace('.', ',')}"
            # final_text += f"{nnf}\t{cfop_map[cfop]}\n"
            final_text += f"{cfop_map[cfop]}\n"
            # print(f"{nnf}\t{cfop_map[cfop]}")
        else:
            # final_text += f"{nnf}\t{cfop_map[cfop]}\n"
            final_text += f"{cfop_map[cfop]}\n"
            # print(f"{nnf}\t{cfop_map[cfop]}")

    win32clipboard.OpenClipboard()
    win32clipboard.EmptyClipboard()
    win32clipboard.SetClipboardText(final_text)
    win32clipboard.CloseClipboard()
    print(final_text)

def romaneio(itens: dict):
    path = "\\\\MARCELO2018\\Documents\\Faturamento\\Refugo e Jato.xlsx"
    wb = openpyxl.load_workbook(path)

    while True:
        tipo_romaneio = input("Refugo ou Jato? ").upper()
        if tipo_romaneio in ["REFUGO", "JATO"]:
            break

    ws = wb[tipo_romaneio]

    # Limpa a planilha antes de preencher
    ws["B2"] = ""
    ws["B3"] = ""
    for i in range(7, 33):
        ws[f"A{i}"] = ""
        ws[f"B{i}"] = ""
        ws[f"D{i}"] = ""
        ws[f"E{i}"] = ""
    ws["B2"] = " / ".join(numero_nfs["MAVIFER"])
    ws["B3"] = " / ".join(numero_nfs["IZAMAC"])

    # Limpa a planilha antes de preencher
    for i in range(7, 33):
        ws[f"A{i}"] = ""
        ws[f"B{i}"] = ""
        ws[f"D{i}"] = ""
        ws[f"E{i}"] = ""

    # itens A7 .. A32
    # quantidade B7 .. B32
    row = 7
    for item in itens["MAVIFER"]:
        ws[f"A{row}"] = item
        ws[f"B{row}"] = itens["MAVIFER"][item]
        row += 1

    # itens D7 .. D32
    # quantidade E7 .. E32
    row = 7
    for item in itens["IZAMAC"]:
        ws[f"D{row}"] = item
        ws[f"E{row}"] = itens["IZAMAC"][item]
        row += 1

    wb.save("\\\\MARCELO2018\\Documents\\Faturamento\\Refugo e Jato.xlsx")


if __name__ == "__main__":
    main()

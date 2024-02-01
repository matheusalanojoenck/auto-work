import csv
from openpyxl import load_workbook

def main():
    paths = "C:\\Users\\Nota Fiscal\\Downloads\\"
    estoque = ["produtos.csv", "produtos (1).csv"]
    recebidos = ["Nota Fiscal Entrada Itens.csv", "Nota Fiscal Entrada Itens (1).csv"]

    itens_recebidos = [{}, {}]
    itens_estoque = [{}, {}]

    # lendo estoque
    for i in range(len(estoque)):
        with open(f"{paths}{estoque[i]}", encoding="ANSI") as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')

            for row in reader:
                item = row["Referencia"]
                qtd = row["Estoque"]
                if item in itens_estoque[i]:
                    itens_estoque[i][item] += qtd
                else:
                    itens_estoque[i][item] = qtd

    # lendo itens recebidos
    for i in range(len(recebidos)):
        with open(f"{paths}{recebidos[i]}") as csvfile:
            for _ in range(3):
                csvfile.__next__()
            reader = list(csv.DictReader(csvfile, delimiter=';'))[:-4]

            for row in reader:
                item = row["Cod. Forn."]
                qtd = int(row["Qtde"])
                if item in itens_recebidos[i]:
                    itens_recebidos[i][item] += qtd
                else:
                    itens_recebidos[i][item] = qtd

    wb = load_workbook(filename="\\\\MARCELO2018\\Documents\\Faturamento\\Estoque PESO - Template.xlsx")
    #wb = load_workbook(filename="C:\\Users\\Nota Fiscal\\Documents\\GitHub\\auto-work\\excel\\Estoque PESO.xlsx")

    index = 2
    for item in itens_estoque[0]:
        wb["MAV - Estq"][f"A{index}"].value = f'=LEFT(B{index},SEARCH("-",B{index},1)-1)'
        wb["MAV - Estq"][f"B{index}"].value = item
        wb["MAV - Estq"][f"C{index}"].value = itens_estoque[0][item]
        wb["MAV - Estq"][f"D{index}"].value = f"=C{index}/VLOOKUP(A{index},POP!A:C,3,FALSE)"
        index += 1

    index = 2
    for item in itens_estoque[1]:
        wb["IZA - Estq"][f"A{index}"].value = f'=LEFT(B{index},SEARCH("-",B{index},1)-1)'
        wb["IZA - Estq"][f"B{index}"].value = item
        wb["IZA - Estq"][f"C{index}"].value = itens_estoque[1][item]
        wb["IZA - Estq"][f"D{index}"].value = f"=C{index}/VLOOKUP(A{index},POP!A:C,3,FALSE)"
        index += 1

    index = 2
    for item in itens_recebidos[0]:
        wb["MAV - Recb"][f"A{index}"].value = f'=LEFT(B{index},SEARCH("-",B{index},1)-1)'
        wb["MAV - Recb"][f"B{index}"].value = item
        wb["MAV - Recb"][f"C{index}"].value = itens_recebidos[0][item]
        wb["MAV - Recb"][f"D{index}"].value = f"=C{index}/VLOOKUP(A{index},POP!A:C,3,FALSE)"
        index += 1

    index = 2
    for item in itens_recebidos[1]:
        wb["IZA - Recb"][f"A{index}"].value = f'=LEFT(B{index},SEARCH("-",B{index},1)-1)'
        wb["IZA - Recb"][f"B{index}"].value = item
        wb["IZA - Recb"][f"C{index}"].value = itens_recebidos[1][item]
        wb["IZA - Recb"][f"D{index}"].value = f"=C{index}/VLOOKUP(A{index},POP!A:C,3,FALSE)"
        index += 1

    wb.save("\\\\MARCELO2018\\Documents\\Faturamento\\Estoque PESO.xlsx")
    #wb.save("C:\\Users\\Nota Fiscal\\Documents\\GitHub\\auto-work\\excel\\Estoque PESO 1.xlsx")

if __name__ == "__main__":
    main()
